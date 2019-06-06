# coding: utf8

import random

import pandas as pd

from openpyxl import Workbook, load_workbook
from pymongo.read_preferences import SecondaryPreferred
# from pymongo.ReadPreference

from tasks.utils.database import databases
from tasks.supplier_statics import Providers, BaseSupplier

SCRIPTURE = databases('scripture', SecondaryPreferred)


def supplier_ctrip_name_coverage(provider):
    service = BaseSupplier(SCRIPTURE)
    table = service.table('hotels', provider)
    hotels_num = table.find().count()

    if not hotels_num:
        return {}
    total = table.find({"ctrip_name": {"$exists": True}}).count()
    coverage = total / hotels_num
    print("供应商%s的中文名覆盖度为：%.2f%% " % (provider.value, coverage * 100))
    return {
        "provider": provider.value,
        "coverage": '%.2f%%' % (coverage * 100)
    }


def city_ctrip_name_coverage(provider):
    title = ['city', 'coverage']
    df = pd.DataFrame(columns=title)

    service = BaseSupplier(SCRIPTURE)
    table = service.table('hotels', provider)
    citys = table.aggregate(
        [
            {
                '$group':
                    {
                        '_id': '$city.name',
                        'num': {'$sum': 1}
                    },
            }
        ]
    )
    for city in citys:
        num = city['num']
        coverage = table.find(
            {
                'city.name': city["_id"],
                'ctrip_name': {
                    '$exists': True
                }
            }
        ).count() / num
        coverage = '%.2f%%' % (coverage * 100)
        print("供应商%s城市%s的中文名覆盖度为：%s" % (provider.value, city["_id"], coverage))
        df.loc[df.shape[0] + 1] = {
            'city': city["_id"],
            'coverage': coverage
        }
    return df


def no_ctrip_name_hotel(provider):
    title = ["provider", "_id", "code", "name"]
    df = pd.DataFrame(columns=title)

    service = BaseSupplier(SCRIPTURE)
    table = service.table('hotels', provider)
    hotels = table.find(
        {
            'ctrip_name': {
                '$exists': False
            }
        }
    )
    # for hotel in hotels:
    #     print("无中文名的酒店：供应商%s酒店%s" % (provider.value, hotel['name']))
    #     df.loc[df.shape[0] + 1] = {
    #         'provider': provider.value,
    #         '_id': str(hotel['_id']),
    #         'code': hotel['code'],
    #         'name': hotel['name']
    #     }
    for hotel in random.sample(hotels, 30):
        df.loc[df.shape[0] + 1] = {
            'provider': provider.value,
            '_id': str(hotel['_id']),
            'code': hotel['code'],
            'name': hotel['name']
        }
    return df


def main():
    path = "./中文名覆盖度.xlsx"
    Workbook().save(path)
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = book

    title = ['provider', 'coverage']
    supplier_df = pd.DataFrame(columns=title)
    for provider in Providers:
        coverage = supplier_ctrip_name_coverage(provider)
        if not coverage:
            continue
        supplier_df.loc[supplier_df.shape[0] + 1] = coverage
        supplier_df.to_excel(
            writer,
            sheet_name="供应商",
            index=False,
            header=True
        )

    for provider in Providers:
        city_ctrip_name_coverage(provider).to_excel(
            writer,
            sheet_name=provider.value,
            index=False,
            header=True
        )

    # dfs = []
    # for provider in Providers:
    #     dfs.append(no_ctrip_name_hotel(provider))
    # no_cname_df = pd.concat(dfs)
    # no_cname_df.to_excel(
    #     writer,
    #     sheet_name="无中文名",
    #     index=False,
    #     header=True
    # )

    writer.save()


if __name__ == '__main__':
    main()
